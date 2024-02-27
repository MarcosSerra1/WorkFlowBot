import openpyxl
import os
from tkinter import messagebox

def salvar_dados_no_excel(dados_json: dict, numero_empresa_desejada: int) -> None:
    '''
    Salva os dados em um arquivo Excel no formato simplificado.

    Parameters:
    - dados_json (dict): Dicionário contendo os dados a serem salvos.
    - numero_empresa_desejada (int): Número da empresa desejada.

    Raises:
    - FileNotFoundError: Se o arquivo de modelo não for encontrado.
    - Exception: Se ocorrer um erro ao salvar os dados no Excel.
    '''

    try:
        # Lista para armazenar os dados dos funcionários da empresa
        dados_funcionarios = []

        # Iterar sobre os dados e encontrar os funcionários da empresa desejada
        for indice, empresa in dados_json['EMPRESA'].items():
            if empresa == numero_empresa_desejada:
                funcionario = {}
                for chave, valor in dados_json.items():
                    funcionario[chave] = valor[indice]
                dados_funcionarios.append(funcionario)

        # Obter o diretório atual do script
        diretorio_atual = os.getcwd() + os.sep

        # Caminho para a planilha modelo
        caminho_relativo = 'doc/evento_simplificado.xlsx'
        # Construir o caminho absoluto para o arquivo Excel
        caminho_absoluto = os.path.join(diretorio_atual, caminho_relativo)

        # Verificar se o arquivo de modelo existe
        if not os.path.isfile(caminho_absoluto):
            raise FileNotFoundError(f'O arquivo "{caminho_absoluto}" não foi encontrado.')

        # Carregar o arquivo Excel de modelo
        planilha = openpyxl.load_workbook(caminho_absoluto)
        # Obtendo a planilha 'evento_simplificado.xls'
        evento_simplificado_sheet = planilha.active

        # Define a linha inicial onde os dados serão colados na planilha de destino
        linha_destino = 2

        # Iterar sobre os dados dos funcionários e copiar para a planilha de destino
        for colaborador in dados_funcionarios:
            if colaborador['ADI'] > 0:
                # Colocando os dados na planilha para adiantamento
                evento_simplificado_sheet.cell(row=linha_destino, column=1, value=colaborador['MATRÍCULA'])
                evento_simplificado_sheet.cell(row=linha_destino, column=2, value=colaborador['COLABORADOR'])
                evento_simplificado_sheet.cell(row=linha_destino, column=3, value='Folha')
                evento_simplificado_sheet.cell(row=linha_destino, column=4, value='503')
                evento_simplificado_sheet.cell(row=linha_destino, column=5, value='Sim')
                evento_simplificado_sheet.cell(row=linha_destino, column=6, value=colaborador['ADI'])
                linha_destino += 1

            if colaborador['COMISSÕES'] > 0:
                # Colocando os dados na planilha para comissões
                evento_simplificado_sheet.cell(row=linha_destino, column=1, value=colaborador['MATRÍCULA'])
                evento_simplificado_sheet.cell(row=linha_destino, column=2, value=colaborador['COLABORADOR'])
                evento_simplificado_sheet.cell(row=linha_destino, column=3, value='Folha')
                evento_simplificado_sheet.cell(row=linha_destino, column=4, value='31')
                evento_simplificado_sheet.cell(row=linha_destino, column=5, value='Sim')
                evento_simplificado_sheet.cell(row=linha_destino, column=6, value=colaborador['COMISSÕES'])
                linha_destino += 1

            if colaborador['COMISSÕES'] > 0:
                # Colocando os dados na planilha para DSR
                evento_simplificado_sheet.cell(row=linha_destino, column=1, value=colaborador['MATRÍCULA'])
                evento_simplificado_sheet.cell(row=linha_destino, column=2, value=colaborador['COLABORADOR'])
                evento_simplificado_sheet.cell(row=linha_destino, column=3, value='Folha')
                evento_simplificado_sheet.cell(row=linha_destino, column=4, value='341')
                evento_simplificado_sheet.cell(row=linha_destino, column=5, value='Não')
                evento_simplificado_sheet.cell(row=linha_destino, column=6, value='0.0')
                linha_destino += 1

        # Salvar o arquivo Excel com um nome específico
        nome_arquivo_excel = f'evento_simplificado_cmr_{numero_empresa_desejada}.xlsx'
        planilha.save(nome_arquivo_excel)

        # Mensagem de sucesso
        messagebox.showinfo('Concluido', f'Planilha salva como {nome_arquivo_excel}')

    except FileNotFoundError as file_error:
        raise FileNotFoundError(f'O arquivo "{caminho_absoluto}" não foi encontrado.') from file_error
    except Exception as e:
        raise Exception('Ocorreu um erro ao salvar os dados no Excel.') from e
